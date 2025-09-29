import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from datetime import datetime
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import glob
from tkcalendar import DateEntry

class GorevFormuApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Delta Proje - Görev Formu Sistemi")
        self.root.geometry("800x600")
        self.root.configure(bg='#f5f5f5')
        
        # Mod: 'menu', 'new', 'edit'
        self.mode = 'menu'
        self.form_data = {}
        self.current_step = 0
        self.form_no = None
        self.is_readonly = False
        
        # Ana frame
        self.main_frame = tk.Frame(root, bg='white', padx=30, pady=30)
        self.main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Ana menüyü göster
        self.show_main_menu()
    
    def get_next_form_no(self):
        """Yeni form numarası al"""
        config_file = 'form_config.json'
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                last_no = config.get('last_form_no', 0)
        else:
            last_no = 0
        
        next_no = last_no + 1
        
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump({'last_form_no': next_no}, f)
        
        return str(next_no).zfill(5)
    
    def get_excel_filename(self, form_no):
        """Form numarasına göre Excel dosya adı"""
        return f"gorev_formu_{form_no}.xlsx"
    
    def load_form_from_excel(self, form_no):
        """Excel dosyasından formu yükle"""
        filename = self.get_excel_filename(form_no)
        if not os.path.exists(filename):
            return None
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            # Excel'den veri oku (basitleştirilmiş - gerçek implementasyon daha detaylı olmalı)
            data = {}
            for row in range(1, ws.max_row + 1):
                key = ws[f'A{row}'].value
                value = ws[f'B{row}'].value
                if key and value:
                    data[key] = value
            
            return data
        except Exception as e:
            messagebox.showerror("Hata", f"Form yüklenemedi: {str(e)}")
            return None
    
    def clear_frame(self):
        """Frame'i temizle"""
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
    def show_main_menu(self):
        """Ana menü ekranı"""
        self.clear_frame()
        self.mode = 'menu'
        
        # Logo/Başlık
        title = tk.Label(
            self.main_frame,
            text="🔧 DELTA PROJE\nGÖREV FORMU SİSTEMİ",
            font=('Arial', 24, 'bold'),
            bg='white',
            fg='#d32f2f'
        )
        title.pack(pady=50)
        
        # Butonlar frame
        button_frame = tk.Frame(self.main_frame, bg='white')
        button_frame.pack(expand=True)
        
        # Yeni Görev Oluştur butonu
        btn_new = tk.Button(
            button_frame,
            text="📝 YENİ GÖREV OLUŞTUR",
            font=('Arial', 16, 'bold'),
            bg='#4dd0e1',
            fg='black',
            width=25,
            height=3,
            command=self.start_new_form,
            cursor='hand2'
        )
        btn_new.pack(pady=15)
        
        # Görev Formu Çağır butonu
        btn_load = tk.Button(
            button_frame,
            text="📂 GÖREV FORMU ÇAĞIR",
            font=('Arial', 16, 'bold'),
            bg='#ffeb3b',
            fg='black',
            width=25,
            height=3,
            command=self.load_existing_form,
            cursor='hand2'
        )
        btn_load.pack(pady=15)
    
    def start_new_form(self):
        """Yeni form oluştur"""
        self.mode = 'new'
        self.form_data = {}
        self.current_step = 0
        self.form_no = self.get_next_form_no()
        self.is_readonly = False
        self.show_step()
    
    def load_existing_form(self):
        """Mevcut formu çağır"""
        self.clear_frame()
        
        # Başlık
        title = tk.Label(
            self.main_frame,
            text="📂 Form Çağır",
            font=('Arial', 20, 'bold'),
            bg='white',
            fg='#d32f2f'
        )
        title.pack(pady=30)
        
        # Açıklama
        info = tk.Label(
            self.main_frame,
            text="Tamamlanacak formun numarasını girin:",
            font=('Arial', 12),
            bg='white'
        )
        info.pack(pady=10)
        
        # Form numarası girişi
        entry_frame = tk.Frame(self.main_frame, bg='white')
        entry_frame.pack(pady=20)
        
        tk.Label(entry_frame, text="Form No:", font=('Arial', 14, 'bold'), bg='white').pack(side='left', padx=10)
        
        form_no_entry = tk.Entry(entry_frame, font=('Arial', 14), width=15, justify='center')
        form_no_entry.pack(side='left', padx=10)
        form_no_entry.focus()
        
        def load_form():
            form_no = form_no_entry.get().strip().zfill(5)
            if not form_no:
                messagebox.showwarning("Uyarı", "Lütfen form numarası girin!")
                return
            
            filename = self.get_excel_filename(form_no)
            if not os.path.exists(filename):
                messagebox.showerror("Hata", f"Form {form_no} bulunamadı!\n\nDosya: {filename}")
                return
            
            # Formu yükle
            self.mode = 'edit'
            self.form_no = form_no
            self.load_partial_form(form_no)
        
        # Butonlar
        btn_frame = tk.Frame(self.main_frame, bg='white')
        btn_frame.pack(pady=30)
        
        tk.Button(
            btn_frame,
            text="✓ FORMU AÇ",
            font=('Arial', 12, 'bold'),
            bg='#4caf50',
            fg='white',
            width=15,
            command=load_form
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="← Geri",
            font=('Arial', 12),
            bg='#ff9800',
            fg='white',
            width=15,
            command=self.show_main_menu
        ).pack(side='left', padx=10)
        
        # Enter tuşu ile aç
        form_no_entry.bind('<Return>', lambda e: load_form())
    
    def load_partial_form(self, form_no):
        """Kısmi dolu formu yükle ve devam et"""
        filename = self.get_excel_filename(form_no)
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            # Basit okuma - gerçek implementasyon daha detaylı mapping gerektirir
            self.form_data = {
                'form_no': form_no,
                'tarih': ws['B2'].value,
                'dok_no': ws['B3'].value,
                'rev_no': ws['B4'].value,
            }
            
            # Görevli personel oku
            for i in range(5):
                cell_value = ws[f'B{6+i}'].value
                if cell_value:
                    self.form_data[f'personel_{i+1}'] = cell_value
            
            # Diğer alanları oku
            row = 12
            if ws[f'A{row}'].value == "Avans Tutarı":
                self.form_data['avans'] = ws[f'B{row}'].value or ''
            row += 1
            if ws[f'A{row}'].value == "Taşeron Şirket":
                self.form_data['taseron'] = ws[f'B{row}'].value or ''
            row += 1
            if ws[f'A{row}'].value == "Görevin Tanımı":
                self.form_data['gorev_tanimi'] = ws[f'B{row}'].value or ''
            row += 1
            if ws[f'A{row}'].value == "Görev Yeri":
                self.form_data['gorev_yeri'] = ws[f'B{row}'].value or ''
            
            # Saat bilgileri boş
            self.current_step = 5  # Saat bilgileri adımından başla
            self.show_step()
            
        except Exception as e:
            messagebox.showerror("Hata", f"Form okunamadı: {str(e)}")
            self.show_main_menu()
    
    def show_step(self):
        """Adımları göster"""
        self.clear_frame()
        
        # Mod kontrolü
        if self.mode == 'new':
            # Yeni form: 0-4 arası adımlar (Görev Yeri'ne kadar)
            if self.current_step > 4:
                self.save_partial_form()
                return
        
        if self.current_step == 0:
            self.step_0_form_bilgileri()
        elif self.current_step == 1:
            self.step_1_gorevli_personel()
        elif self.current_step == 2:
            self.step_2_avans_taseron()
        elif self.current_step == 3:
            self.step_3_gorev_tanimi()
        elif self.current_step == 4:
            self.step_4_gorev_yeri()
        elif self.current_step == 5:
            self.step_5_saat_bilgileri()
        elif self.current_step == 6:
            self.step_6_arac_bilgisi()
        elif self.current_step == 7:
            self.step_7_hazirlayan()
        elif self.current_step == 8:
            self.show_summary()
    
    def step_0_form_bilgileri(self):
        """Adım 0: Form bilgileri"""
        readonly = self.mode == 'edit'
        
        tk.Label(self.main_frame, text="📋 Form Bilgileri", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        form_frame = tk.Frame(self.main_frame, bg='white')
        form_frame.pack(pady=20)
        
        # Form No
        tk.Label(form_frame, text="Form No:", font=('Arial', 12, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=10)
        form_no_label = tk.Label(form_frame, text=self.form_no, font=('Arial', 12), bg='#e3f2fd', width=20, anchor='w')
        form_no_label.grid(row=0, column=1, padx=10, pady=10)
        
        # Tarih
        tk.Label(form_frame, text="Tarih:", font=('Arial', 12, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=10)
        tarih_value = self.form_data.get('tarih', datetime.now().strftime('%d.%m.%Y'))
        tarih_label = tk.Label(form_frame, text=tarih_value, font=('Arial', 12), bg='#e3f2fd', width=20, anchor='w')
        tarih_label.grid(row=1, column=1, padx=10, pady=10)
        
        # DOK.NO
        tk.Label(form_frame, text="DOK.NO:", font=('Arial', 12, 'bold'), bg='white').grid(row=2, column=0, sticky='w', pady=10)
        dok_entry = tk.Entry(form_frame, font=('Arial', 12), width=20)
        dok_entry.insert(0, self.form_data.get('dok_no', 'F-001'))
        dok_entry.grid(row=2, column=1, padx=10, pady=10)
        if readonly:
            dok_entry.config(state='readonly', bg='#f0f0f0')
        self.form_data['dok_no_widget'] = dok_entry
        
        # REV.NO/TRH
        tk.Label(form_frame, text="REV.NO/TRH:", font=('Arial', 12, 'bold'), bg='white').grid(row=3, column=0, sticky='w', pady=10)
        rev_entry = tk.Entry(form_frame, font=('Arial', 12), width=20)
        rev_entry.insert(0, self.form_data.get('rev_no', ''))
        rev_entry.grid(row=3, column=1, padx=10, pady=10)
        if readonly:
            rev_entry.config(state='readonly', bg='#f0f0f0')
        self.form_data['rev_no_widget'] = rev_entry
        
        self.form_data['tarih'] = tarih_value
        
        self.add_navigation_buttons(readonly)
    
    def step_1_gorevli_personel(self):
        """Adım 1: Görevli personel"""
        readonly = self.mode == 'edit'
        
        tk.Label(self.main_frame, text="👥 Görevli Personel", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        personel_options = [
            "Ahmet Yılmaz", "Mehmet Demir", "Ali Kaya", "Veli Çelik",
            "Hasan Şahin", "Hüseyin Aydın", "İbrahim Özdemir", "Mustafa Arslan",
            "Emre Doğan", "Burak Yıldız"
        ]
        
        form_frame = tk.Frame(self.main_frame, bg='white')
        form_frame.pack(pady=20)
        
        self.form_data['personel_widgets'] = []
        
        for i in range(5):
            tk.Label(form_frame, text=f"Personel {i+1}:", font=('Arial', 12, 'bold'), bg='white').grid(row=i, column=0, sticky='w', pady=10, padx=10)
            
            if readonly:
                value = self.form_data.get(f'personel_{i+1}', '')
                label = tk.Label(form_frame, text=value, font=('Arial', 12), bg='#f0f0f0', width=25, anchor='w')
                label.grid(row=i, column=1, padx=10, pady=10)
                self.form_data['personel_widgets'].append(label)
            else:
                combo = ttk.Combobox(form_frame, values=personel_options, font=('Arial', 12), width=23, state='readonly')
                combo.set(self.form_data.get(f'personel_{i+1}', ''))
                combo.grid(row=i, column=1, padx=10, pady=10)
                self.form_data['personel_widgets'].append(combo)
        
        self.add_navigation_buttons(readonly)
    
    def step_2_avans_taseron(self):
        """Adım 2: Avans ve Taşeron"""
        readonly = self.mode == 'edit'
        
        tk.Label(self.main_frame, text="💰 Avans ve Taşeron Bilgileri", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        form_frame = tk.Frame(self.main_frame, bg='white')
        form_frame.pack(pady=40)
        
        # Avans
        tk.Label(form_frame, text="Avans Tutarı:", font=('Arial', 12, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=15)
        avans_entry = tk.Entry(form_frame, font=('Arial', 12), width=30)
        avans_entry.insert(0, self.form_data.get('avans', ''))
        avans_entry.grid(row=0, column=1, padx=10, pady=15)
        if readonly:
            avans_entry.config(state='readonly', bg='#f0f0f0')
        self.form_data['avans_widget'] = avans_entry
        
        # Taşeron
        tk.Label(form_frame, text="Taşeron Şirket:", font=('Arial', 12, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=15)
        
        taseron_options = ["Yok", "ABC İnşaat", "XYZ Teknik", "Marmara Mühendislik", "Anadolu Yapı"]
        
        if readonly:
            value = self.form_data.get('taseron', '')
            label = tk.Label(form_frame, text=value, font=('Arial', 12), bg='#f0f0f0', width=28, anchor='w')
            label.grid(row=1, column=1, padx=10, pady=15)
            self.form_data['taseron_widget'] = label
        else:
            taseron_combo = ttk.Combobox(form_frame, values=taseron_options, font=('Arial', 12), width=28)
            taseron_combo.set(self.form_data.get('taseron', ''))
            taseron_combo.grid(row=1, column=1, padx=10, pady=15)
            self.form_data['taseron_widget'] = taseron_combo
        
        self.add_navigation_buttons(readonly)
    
    def step_3_gorev_tanimi(self):
        """Adım 3: Görev Tanımı"""
        readonly = self.mode == 'edit'
        
        tk.Label(self.main_frame, text="📝 Görevin Tanımı", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        text_widget = scrolledtext.ScrolledText(self.main_frame, font=('Arial', 11), width=70, height=15, wrap='word')
        text_widget.pack(pady=20, padx=20)
        text_widget.insert('1.0', self.form_data.get('gorev_tanimi', ''))
        
        if readonly:
            text_widget.config(state='disabled', bg='#f0f0f0')
        
        self.form_data['gorev_tanimi_widget'] = text_widget
        
        self.add_navigation_buttons(readonly)
    
    def step_4_gorev_yeri(self):
        """Adım 4: Görev Yeri"""
        readonly = self.mode == 'edit'
        
        tk.Label(self.main_frame, text="📍 Görev Yeri", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        text_widget = scrolledtext.ScrolledText(self.main_frame, font=('Arial', 11), width=70, height=15, wrap='word')
        text_widget.pack(pady=20, padx=20)
        text_widget.insert('1.0', self.form_data.get('gorev_yeri', ''))
        
        if readonly:
            text_widget.config(state='disabled', bg='#f0f0f0')
        
        self.form_data['gorev_yeri_widget'] = text_widget
        
        self.add_navigation_buttons(readonly)
    
    def step_5_saat_bilgileri(self):
        """Adım 5: Saat bilgileri"""
        tk.Label(self.main_frame, text="🕐 Saat ve Tarih Bilgileri", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        # Scroll frame
        canvas = tk.Canvas(self.main_frame, bg='white', highlightthickness=0)
        scrollbar = tk.Scrollbar(self.main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        form_frame = tk.Frame(scrollable_frame, bg='white')
        form_frame.pack(pady=10, padx=20)
        
        row = 0
        
        # Yola Çıkış
        tk.Label(form_frame, text="Yola Çıkış:", font=('Arial', 12, 'bold'), bg='white').grid(row=row, column=0, sticky='w', pady=10)
        tk.Label(form_frame, text="Tarih:", bg='white').grid(row=row, column=1, sticky='e', padx=5)
        yola_cikis_tarih = DateEntry(form_frame, font=('Arial', 11), width=12, background='#4dd0e1', 
                                      foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy',
                                      locale='tr_TR')
        yola_cikis_tarih.grid(row=row, column=2, padx=5)
        if self.form_data.get('yola_cikis_tarih'):
            try:
                yola_cikis_tarih.set_date(datetime.strptime(self.form_data.get('yola_cikis_tarih'), '%d.%m.%Y'))
            except:
                pass
        
        tk.Label(form_frame, text="Saat:", bg='white').grid(row=row, column=3, sticky='e', padx=5)
        
        # Saat frame
        saat_frame1 = tk.Frame(form_frame, bg='white')
        saat_frame1.grid(row=row, column=4, padx=5)
        
        yola_cikis_saat = ttk.Spinbox(saat_frame1, from_=0, to=23, width=3, format='%02.0f', font=('Arial', 11))
        yola_cikis_saat.pack(side='left')
        tk.Label(saat_frame1, text=":", bg='white', font=('Arial', 11, 'bold')).pack(side='left')
        yola_cikis_dakika = ttk.Spinbox(saat_frame1, from_=0, to=59, width=3, format='%02.0f', font=('Arial', 11))
        yola_cikis_dakika.pack(side='left')
        
        if self.form_data.get('yola_cikis_saat'):
            try:
                h, m = self.form_data.get('yola_cikis_saat', '00:00').split(':')
                yola_cikis_saat.set(int(h))
                yola_cikis_dakika.set(int(m))
            except:
                pass
        
        row += 1
        
        # Dönüş
        tk.Label(form_frame, text="Dönüş:", font=('Arial', 12, 'bold'), bg='white').grid(row=row, column=0, sticky='w', pady=10)
        tk.Label(form_frame, text="Tarih:", bg='white').grid(row=row, column=1, sticky='e', padx=5)
        donus_tarih = DateEntry(form_frame, font=('Arial', 11), width=12, background='#4dd0e1',
                               foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy',
                               locale='tr_TR')
        donus_tarih.grid(row=row, column=2, padx=5)
        if self.form_data.get('donus_tarih'):
            try:
                donus_tarih.set_date(datetime.strptime(self.form_data.get('donus_tarih'), '%d.%m.%Y'))
            except:
                pass
        
        tk.Label(form_frame, text="Saat:", bg='white').grid(row=row, column=3, sticky='e', padx=5)
        
        saat_frame2 = tk.Frame(form_frame, bg='white')
        saat_frame2.grid(row=row, column=4, padx=5)
        
        donus_saat = ttk.Spinbox(saat_frame2, from_=0, to=23, width=3, format='%02.0f', font=('Arial', 11))
        donus_saat.pack(side='left')
        tk.Label(saat_frame2, text=":", bg='white', font=('Arial', 11, 'bold')).pack(side='left')
        donus_dakika = ttk.Spinbox(saat_frame2, from_=0, to=59, width=3, format='%02.0f', font=('Arial', 11))
        donus_dakika.pack(side='left')
        
        if self.form_data.get('donus_saat'):
            try:
                h, m = self.form_data.get('donus_saat', '00:00').split(':')
                donus_saat.set(int(h))
                donus_dakika.set(int(m))
            except:
                pass
        
        row += 1
        
        # Çalışma Başlangıç
        tk.Label(form_frame, text="Çalışma Başlangıç:", font=('Arial', 12, 'bold'), bg='white').grid(row=row, column=0, sticky='w', pady=10)
        tk.Label(form_frame, text="Tarih:", bg='white').grid(row=row, column=1, sticky='e', padx=5)
        calisma_baslangic_tarih = DateEntry(form_frame, font=('Arial', 11), width=12, background='#4dd0e1',
                                           foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy',
                                           locale='tr_TR')
        calisma_baslangic_tarih.grid(row=row, column=2, padx=5)
        if self.form_data.get('calisma_baslangic_tarih'):
            try:
                calisma_baslangic_tarih.set_date(datetime.strptime(self.form_data.get('calisma_baslangic_tarih'), '%d.%m.%Y'))
            except:
                pass
        
        tk.Label(form_frame, text="Saat:", bg='white').grid(row=row, column=3, sticky='e', padx=5)
        
        saat_frame3 = tk.Frame(form_frame, bg='white')
        saat_frame3.grid(row=row, column=4, padx=5)
        
        calisma_baslangic_saat = ttk.Spinbox(saat_frame3, from_=0, to=23, width=3, format='%02.0f', font=('Arial', 11))
        calisma_baslangic_saat.pack(side='left')
        tk.Label(saat_frame3, text=":", bg='white', font=('Arial', 11, 'bold')).pack(side='left')
        calisma_baslangic_dakika = ttk.Spinbox(saat_frame3, from_=0, to=59, width=3, format='%02.0f', font=('Arial', 11))
        calisma_baslangic_dakika.pack(side='left')
        
        if self.form_data.get('calisma_baslangic_saat'):
            try:
                h, m = self.form_data.get('calisma_baslangic_saat', '00:00').split(':')
                calisma_baslangic_saat.set(int(h))
                calisma_baslangic_dakika.set(int(m))
            except:
                pass
        
        row += 1
        
        # Çalışma Bitiş
        tk.Label(form_frame, text="Çalışma Bitiş:", font=('Arial', 12, 'bold'), bg='white').grid(row=row, column=0, sticky='w', pady=10)
        tk.Label(form_frame, text="Tarih:", bg='white').grid(row=row, column=1, sticky='e', padx=5)
        calisma_bitis_tarih = DateEntry(form_frame, font=('Arial', 11), width=12, background='#4dd0e1',
                                        foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy',
                                        locale='tr_TR')
        calisma_bitis_tarih.grid(row=row, column=2, padx=5)
        if self.form_data.get('calisma_bitis_tarih'):
            try:
                calisma_bitis_tarih.set_date(datetime.strptime(self.form_data.get('calisma_bitis_tarih'), '%d.%m.%Y'))
            except:
                pass
        
        tk.Label(form_frame, text="Saat:", bg='white').grid(row=row, column=3, sticky='e', padx=5)
        
        saat_frame4 = tk.Frame(form_frame, bg='white')
        saat_frame4.grid(row=row, column=4, padx=5)
        
        calisma_bitis_saat = ttk.Spinbox(saat_frame4, from_=0, to=23, width=3, format='%02.0f', font=('Arial', 11))
        calisma_bitis_saat.pack(side='left')
        tk.Label(saat_frame4, text=":", bg='white', font=('Arial', 11, 'bold')).pack(side='left')
        calisma_bitis_dakika = ttk.Spinbox(saat_frame4, from_=0, to=59, width=3, format='%02.0f', font=('Arial', 11))
        calisma_bitis_dakika.pack(side='left')
        
        if self.form_data.get('calisma_bitis_saat'):
            try:
                h, m = self.form_data.get('calisma_bitis_saat', '00:00').split(':')
                calisma_bitis_saat.set(int(h))
                calisma_bitis_dakika.set(int(m))
            except:
                pass
        
        row += 1
        
        # Mola Süresi
        tk.Label(form_frame, text="Toplam Mola Süresi:", font=('Arial', 12, 'bold'), bg='white').grid(row=row, column=0, sticky='w', pady=10)
        mola_suresi = ttk.Spinbox(form_frame, from_=0, to=480, width=10, font=('Arial', 11))
        mola_suresi.set(self.form_data.get('mola_suresi', '0'))
        mola_suresi.grid(row=row, column=2, padx=5)
        tk.Label(form_frame, text="dakika", bg='white').grid(row=row, column=3, sticky='w', padx=5)
        
        # Widget'ları sakla
        self.form_data['yola_cikis_tarih_widget'] = yola_cikis_tarih
        self.form_data['yola_cikis_saat_widget'] = yola_cikis_saat
        self.form_data['yola_cikis_dakika_widget'] = yola_cikis_dakika
        self.form_data['donus_tarih_widget'] = donus_tarih
        self.form_data['donus_saat_widget'] = donus_saat
        self.form_data['donus_dakika_widget'] = donus_dakika
        self.form_data['calisma_baslangic_tarih_widget'] = calisma_baslangic_tarih
        self.form_data['calisma_baslangic_saat_widget'] = calisma_baslangic_saat
        self.form_data['calisma_baslangic_dakika_widget'] = calisma_baslangic_dakika
        self.form_data['calisma_bitis_tarih_widget'] = calisma_bitis_tarih
        self.form_data['calisma_bitis_saat_widget'] = calisma_bitis_saat
        self.form_data['calisma_bitis_dakika_widget'] = calisma_bitis_dakika
        self.form_data['mola_suresi_widget'] = mola_suresi
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.add_navigation_buttons(False, canvas_parent=True)
    
    def step_6_arac_bilgisi(self):
        """Adım 6: Araç bilgisi"""
        tk.Label(self.main_frame, text="🚗 Araç Bilgisi", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        form_frame = tk.Frame(self.main_frame, bg='white')
        form_frame.pack(pady=40)
        
        tk.Label(form_frame, text="Araç Plaka No:", font=('Arial', 12, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=15)
        
        arac_options = [
            "34 ABC 123", "06 DEF 456", "41 GHI 789",
            "16 JKL 012", "35 MNO 345"
        ]
        
        arac_combo = ttk.Combobox(form_frame, values=arac_options, font=('Arial', 12), width=28, state='readonly')
        arac_combo.set(self.form_data.get('arac_plaka', ''))
        arac_combo.grid(row=0, column=1, padx=10, pady=15)
        self.form_data['arac_plaka_widget'] = arac_combo
        
        self.add_navigation_buttons(False)
    
    def step_7_hazirlayan(self):
        """Adım 7: Hazırlayan"""
        tk.Label(self.main_frame, text="✍️ Hazırlayan / Görevlendiren", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        form_frame = tk.Frame(self.main_frame, bg='white')
        form_frame.pack(pady=40)
        
        tk.Label(form_frame, text="Ad Soyad:", font=('Arial', 12, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=15)
        
        hazirlayan_options = [
            "Ahmet Yılmaz", "Mehmet Demir", "Ali Kaya",
            "Veli Çelik", "Hasan Şahin"
        ]
        
        hazirlayan_combo = ttk.Combobox(form_frame, values=hazirlayan_options, font=('Arial', 12), width=28, state='readonly')
        hazirlayan_combo.set(self.form_data.get('hazirlayan', ''))
        hazirlayan_combo.grid(row=0, column=1, padx=10, pady=15)
        self.form_data['hazirlayan_widget'] = hazirlayan_combo
        
        self.add_navigation_buttons(False)
    
    def show_summary(self):
        """Özet ekranı"""
        tk.Label(self.main_frame, text="📊 Form Özeti", font=('Arial', 18, 'bold'), bg='white', fg='#d32f2f').pack(pady=20)
        
        # Scroll text
        summary_text = scrolledtext.ScrolledText(self.main_frame, font=('Arial', 10), width=80, height=20, wrap='word')
        summary_text.pack(pady=20, padx=20)
        
        # Özet oluştur
        summary = f"""
═══════════════════════════════════════════════════════
                    GÖREV FORMU ÖZETİ
═══════════════════════════════════════════════════════

Form No: {self.form_no}
Tarih: {self.form_data.get('tarih', '')}
DOK.NO: {self.form_data.get('dok_no', '')}
REV.NO/TRH: {self.form_data.get('rev_no', '')}

───────────────────────────────────────────────────────
GÖREVLİ PERSONEL
───────────────────────────────────────────────────────
"""
        
        for i in range(5):
            personel = self.form_data.get(f'personel_{i+1}', '')
            if personel:
                summary += f"{i+1}. {personel}\n"
        
        summary += f"""
───────────────────────────────────────────────────────
MALİ BİLGİLER
───────────────────────────────────────────────────────
Avans Tutarı: {self.form_data.get('avans', '')}
Taşeron Şirket: {self.form_data.get('taseron', '')}

───────────────────────────────────────────────────────
GÖREV DETAYLARI
───────────────────────────────────────────────────────
Görevin Tanımı:
{self.form_data.get('gorev_tanimi', '')}

Görev Yeri:
{self.form_data.get('gorev_yeri', '')}

───────────────────────────────────────────────────────
ZAMAN BİLGİLERİ
───────────────────────────────────────────────────────
Yola Çıkış: {self.form_data.get('yola_cikis_tarih', '')} {self.form_data.get('yola_cikis_saat', '')}
Dönüş: {self.form_data.get('donus_tarih', '')} {self.form_data.get('donus_saat', '')}
Çalışma Başlangıç: {self.form_data.get('calisma_baslangic_tarih', '')} {self.form_data.get('calisma_baslangic_saat', '')}
Çalışma Bitiş: {self.form_data.get('calisma_bitis_tarih', '')} {self.form_data.get('calisma_bitis_saat', '')}
Toplam Mola: {self.form_data.get('mola_suresi', '')} dakika

───────────────────────────────────────────────────────
DİĞER BİLGİLER
───────────────────────────────────────────────────────
Araç Plaka: {self.form_data.get('arac_plaka', '')}
Hazırlayan: {self.form_data.get('hazirlayan', '')}

═══════════════════════════════════════════════════════
"""
        
        summary_text.insert('1.0', summary)
        summary_text.config(state='disabled')
        
        # Butonlar
        btn_frame = tk.Frame(self.main_frame, bg='white')
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="💾 KAYDET",
            font=('Arial', 14, 'bold'),
            bg='#4caf50',
            fg='white',
            width=15,
            height=2,
            command=self.save_form
        ).pack(side='left', padx=10)
        
        tk.Button(
            btn_frame,
            text="← Geri",
            font=('Arial', 12),
            bg='#ff9800',
            fg='white',
            width=15,
            command=self.previous_step
        ).pack(side='left', padx=10)
    
    def collect_form_data(self):
        """Widget'lardan veri topla"""
        try:
            # Form bilgileri
            if 'dok_no_widget' in self.form_data:
                self.form_data['dok_no'] = self.form_data['dok_no_widget'].get()
            if 'rev_no_widget' in self.form_data:
                self.form_data['rev_no'] = self.form_data['rev_no_widget'].get()
            
            # Personel
            if 'personel_widgets' in self.form_data:
                for i, widget in enumerate(self.form_data['personel_widgets']):
                    if isinstance(widget, ttk.Combobox):
                        self.form_data[f'personel_{i+1}'] = widget.get()
                    elif isinstance(widget, tk.Label):
                        # Readonly mode - zaten form_data'da var
                        pass
            
            # Avans ve Taşeron
            if 'avans_widget' in self.form_data:
                if isinstance(self.form_data['avans_widget'], tk.Entry):
                    self.form_data['avans'] = self.form_data['avans_widget'].get()
            if 'taseron_widget' in self.form_data:
                if isinstance(self.form_data['taseron_widget'], ttk.Combobox):
                    self.form_data['taseron'] = self.form_data['taseron_widget'].get()
            
            # Görev tanımı ve yeri
            if 'gorev_tanimi_widget' in self.form_data:
                self.form_data['gorev_tanimi'] = self.form_data['gorev_tanimi_widget'].get('1.0', 'end-1c')
            if 'gorev_yeri_widget' in self.form_data:
                self.form_data['gorev_yeri'] = self.form_data['gorev_yeri_widget'].get('1.0', 'end-1c')
            
            # Saat bilgileri
            if 'yola_cikis_tarih_widget' in self.form_data:
                self.form_data['yola_cikis_tarih'] = self.form_data['yola_cikis_tarih_widget'].get_date().strftime('%d.%m.%Y')
                h = self.form_data['yola_cikis_saat_widget'].get()
                m = self.form_data['yola_cikis_dakika_widget'].get()
                self.form_data['yola_cikis_saat'] = f"{int(h):02d}:{int(m):02d}"
                
                self.form_data['donus_tarih'] = self.form_data['donus_tarih_widget'].get_date().strftime('%d.%m.%Y')
                h = self.form_data['donus_saat_widget'].get()
                m = self.form_data['donus_dakika_widget'].get()
                self.form_data['donus_saat'] = f"{int(h):02d}:{int(m):02d}"
                
                self.form_data['calisma_baslangic_tarih'] = self.form_data['calisma_baslangic_tarih_widget'].get_date().strftime('%d.%m.%Y')
                h = self.form_data['calisma_baslangic_saat_widget'].get()
                m = self.form_data['calisma_baslangic_dakika_widget'].get()
                self.form_data['calisma_baslangic_saat'] = f"{int(h):02d}:{int(m):02d}"
                
                self.form_data['calisma_bitis_tarih'] = self.form_data['calisma_bitis_tarih_widget'].get_date().strftime('%d.%m.%Y')
                h = self.form_data['calisma_bitis_saat_widget'].get()
                m = self.form_data['calisma_bitis_dakika_widget'].get()
                self.form_data['calisma_bitis_saat'] = f"{int(h):02d}:{int(m):02d}"
                
                self.form_data['mola_suresi'] = self.form_data['mola_suresi_widget'].get()
            
            # Araç ve hazırlayan
            if 'arac_plaka_widget' in self.form_data:
                self.form_data['arac_plaka'] = self.form_data['arac_plaka_widget'].get()
            if 'hazirlayan_widget' in self.form_data:
                self.form_data['hazirlayan'] = self.form_data['hazirlayan_widget'].get()
        except Exception as e:
            # Hata durumunda sessizce devam et
            pass
    
    def save_partial_form(self):
        """Kısmi formu kaydet (Görev Yeri'ne kadar)"""
        self.collect_form_data()
        
        filename = self.get_excel_filename(self.form_no)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Görev Formu"
            
            # Stil
            header_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Başlık
            ws[f'A{row}'] = "DELTA PROJE - GÖREV FORMU"
            ws[f'A{row}'].font = Font(size=16, bold=True, color='D32F2F')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Form bilgileri
            ws[f'A{row}'] = "Form No"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_no
            row += 1
            
            ws[f'A{row}'] = "Tarih"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('tarih', '')
            row += 1
            
            ws[f'A{row}'] = "DOK.NO"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('dok_no', '')
            row += 1
            
            ws[f'A{row}'] = "REV.NO/TRH"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('rev_no', '')
            row += 1
            
            # Personel
            ws[f'A{row}'] = "Görevli Personel"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            for i in range(5):
                ws[f'A{row}'] = f"Personel {i+1}"
                ws[f'B{row}'] = self.form_data.get(f'personel_{i+1}', '')
                row += 1
            
            # Diğer bilgiler
            ws[f'A{row}'] = "Avans Tutarı"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('avans', '')
            row += 1
            
            ws[f'A{row}'] = "Taşeron Şirket"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('taseron', '')
            row += 1
            
            ws[f'A{row}'] = "Görevin Tanımı"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('gorev_tanimi', '')
            row += 1
            
            ws[f'A{row}'] = "Görev Yeri"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = self.form_data.get('gorev_yeri', '')
            row += 1
            
            # Durum
            ws[f'A{row}'] = "DURUM"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
            ws[f'B{row}'] = "YARIM"
            ws[f'B{row}'].fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 60
            
            wb.save(filename)
            
            messagebox.showinfo(
                "Başarılı",
                f"Form oluşturuldu!\n\nForm No: {self.form_no}\nDosya: {filename}\n\nGörev tamamlandığında 'GÖREV FORMU ÇAĞIR' ile bu formu açıp kalan kısımları doldurun."
            )
            
            self.show_main_menu()
            
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası: {str(e)}")
    
    def save_form(self):
        """Tam formu kaydet"""
        self.collect_form_data()
        
        filename = self.get_excel_filename(self.form_no)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Görev Formu"
            
            # Stil
            header_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
            
            row = 1
            
            # Başlık
            ws[f'A{row}'] = "DELTA PROJE - GÖREV FORMU"
            ws[f'A{row}'].font = Font(size=16, bold=True, color='D32F2F')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Tüm bilgileri yaz
            data_map = [
                ("Form No", self.form_no),
                ("Tarih", self.form_data.get('tarih', '')),
                ("DOK.NO", self.form_data.get('dok_no', '')),
                ("REV.NO/TRH", self.form_data.get('rev_no', '')),
                ("", ""),
                ("Görevli Personel", ""),
            ]
            
            for label, value in data_map:
                if label:
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'A{row}'].fill = header_fill
                    ws[f'B{row}'] = value
                row += 1
            
            # Personel listesi
            for i in range(5):
                ws[f'A{row}'] = f"Personel {i+1}"
                ws[f'B{row}'] = self.form_data.get(f'personel_{i+1}', '')
                row += 1
            
            row += 1
            
            # Diğer tüm alanlar
            all_data = [
                ("Avans Tutarı", self.form_data.get('avans', '')),
                ("Taşeron Şirket", self.form_data.get('taseron', '')),
                ("Görevin Tanımı", self.form_data.get('gorev_tanimi', '')),
                ("Görev Yeri", self.form_data.get('gorev_yeri', '')),
                ("", ""),
                ("Yola Çıkış", f"{self.form_data.get('yola_cikis_tarih', '')} {self.form_data.get('yola_cikis_saat', '')}"),
                ("Dönüş", f"{self.form_data.get('donus_tarih', '')} {self.form_data.get('donus_saat', '')}"),
                ("Çalışma Başlangıç", f"{self.form_data.get('calisma_baslangic_tarih', '')} {self.form_data.get('calisma_baslangic_saat', '')}"),
                ("Çalışma Bitiş", f"{self.form_data.get('calisma_bitis_tarih', '')} {self.form_data.get('calisma_bitis_saat', '')}"),
                ("Toplam Mola", f"{self.form_data.get('mola_suresi', '')} dakika"),
                ("", ""),
                ("Araç Plaka No", self.form_data.get('arac_plaka', '')),
                ("Hazırlayan", self.form_data.get('hazirlayan', '')),
                ("", ""),
                ("DURUM", "TAMAMLANDI"),
            ]
            
            for label, value in all_data:
                if label:
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = Font(bold=True)
                    if label == "DURUM":
                        ws[f'A{row}'].fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                        ws[f'B{row}'].fill = PatternFill(start_color='81C784', end_color='81C784', fill_type='solid')
                    else:
                        ws[f'A{row}'].fill = header_fill
                    ws[f'B{row}'] = value
                row += 1
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 60
            
            wb.save(filename)
            
            messagebox.showinfo(
                "Başarılı",
                f"Form başarıyla tamamlandı ve kaydedildi!\n\nForm No: {self.form_no}\nDosya: {filename}"
            )
            
            self.show_main_menu()
            
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası: {str(e)}")
    
    def add_navigation_buttons(self, readonly=False, canvas_parent=False):
        """Navigasyon butonları ekle"""
        parent = self.main_frame if not canvas_parent else self.root
        
        btn_frame = tk.Frame(parent, bg='white')
        if canvas_parent:
            btn_frame.pack(side='bottom', pady=10)
        else:
            btn_frame.pack(side='bottom', pady=20)
        
        if self.current_step > 0:
            tk.Button(
                btn_frame,
                text="← Geri",
                font=('Arial', 12),
                bg='#ff9800',
                fg='white',
                width=12,
                command=self.previous_step
            ).pack(side='left', padx=10)
        
        if self.mode == 'new' and self.current_step >= 4:
            # Yeni form modunda Görev Yeri'nden sonra kaydet
            tk.Button(
                btn_frame,
                text="💾 Kaydet",
                font=('Arial', 12, 'bold'),
                bg='#4caf50',
                fg='white',
                width=12,
                command=lambda: self.next_step(save_partial=True)
            ).pack(side='left', padx=10)
        else:
            # Normal ilerleme
            tk.Button(
                btn_frame,
                text="İleri →",
                font=('Arial', 12, 'bold'),
                bg='#4dd0e1',
                fg='black',
                width=12,
                command=self.next_step
            ).pack(side='left', padx=10)
    
    def next_step(self, save_partial=False):
        """Sonraki adım"""
        self.collect_form_data()
        
        if save_partial:
            self.save_partial_form()
            return
        
        self.current_step += 1
        self.show_step()
    
    def previous_step(self):
        """Önceki adım"""
        self.collect_form_data()
        
        if self.mode == 'edit' and self.current_step == 5:
            # Edit modunda geri dönmeye izin verme
            messagebox.showwarning("Uyarı", "Önceki adımlara dönüş yapılamaz!")
            return
        
        self.current_step -= 1
        self.show_step()


if __name__ == "__main__":
    root = tk.Tk()
    app = GorevFormuApp(root)
    root.mainloop()
