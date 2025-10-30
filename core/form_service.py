# -*- coding: utf-8 -*-
"""Servis katmanı: Görev formu veri işlemleri."""
from __future__ import annotations

import io
import json
import os
import sqlite3
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas


DB_FILENAME = "forms.db"

PERSONEL_FIELDS: Tuple[str, ...] = tuple(f"personel_{index}" for index in range(1, 6))


class FormServiceError(Exception):
    """Servis katmanına özgü hata sınıfı."""


@dataclass
class FormStatus:
    """Form durumunu ve eksik alanları temsil eder."""

    code: str
    missing_fields: List[str]

    @property
    def is_complete(self) -> bool:
        return self.code.upper() == "TAMAMLANDI"


def get_db_path(base_path: str = ".") -> str:
    """Veritabanı dosya yolunu döndür."""

    return os.path.join(base_path, DB_FILENAME)


def get_connection(base_path: str = ".") -> sqlite3.Connection:
    """Uygulama genelinde kullanılan veritabanı bağlantısını döndür."""

    return _connect(base_path)


def _connect(base_path: str = ".") -> sqlite3.Connection:
    db_path = get_db_path(base_path)
    directory = os.path.dirname(db_path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    _ensure_schema(connection)
    return connection


def _ensure_schema(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            password_hash TEXT,
            role TEXT NOT NULL,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_no TEXT NOT NULL UNIQUE,
            tarih TEXT,
            tarih_iso TEXT,
            dok_no TEXT,
            rev_no TEXT,
            avans TEXT,
            taseron TEXT,
            gorev_tanimi TEXT,
            gorev_yeri TEXT,
            gorev_yeri_lower TEXT,
            gorev_il TEXT,
            gorev_ilce TEXT,
            gorev_firma TEXT,
            gorev_tarih TEXT,
            gorev_tarih_iso TEXT,
            yapilan_isler TEXT,
            gorev_ekleri TEXT,
            harcama_bildirimleri TEXT,
            yola_cikis_tarih TEXT,
            yola_cikis_tarih_iso TEXT,
            yola_cikis_saat TEXT,
            donus_tarih TEXT,
            donus_tarih_iso TEXT,
            donus_saat TEXT,
            calisma_baslangic_tarih TEXT,
            calisma_baslangic_tarih_iso TEXT,
            calisma_baslangic_saat TEXT,
            calisma_bitis_tarih TEXT,
            calisma_bitis_tarih_iso TEXT,
            calisma_bitis_saat TEXT,
            mola_suresi TEXT,
            arac_plaka TEXT,
            hazirlayan TEXT,
            durum TEXT,
            personel_1 TEXT,
            personel_2 TEXT,
            personel_3 TEXT,
            personel_4 TEXT,
            personel_5 TEXT,
            personel_search TEXT,
            last_step INTEGER DEFAULT 0,
            assigned_to_user_id INTEGER,
            assigned_by_user_id INTEGER,
            assigned_at TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(assigned_to_user_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY(assigned_by_user_id) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    connection.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_forms_form_no
            ON forms (form_no)
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS form_sequence (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            last_no INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    connection.execute(
        "INSERT OR IGNORE INTO form_sequence (id, last_no) VALUES (1, 0)"
    )

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS task_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_name TEXT NOT NULL,
            customer_phone TEXT,
            customer_email TEXT,
            customer_address TEXT,
            request_description TEXT NOT NULL,
            requirements TEXT,
            urgency TEXT DEFAULT 'normal',
            requested_by_user_id INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            notes TEXT,
            assigned_to_user_id INTEGER,
            converted_form_no TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(requested_by_user_id) REFERENCES users(id),
            FOREIGN KEY(assigned_to_user_id) REFERENCES users(id)
        )
        """
    )

    existing_columns = {
        row["name"] for row in connection.execute("PRAGMA table_info(forms)")
    }
    for column, definition in (
        ("yapilan_isler", "TEXT"),
        ("gorev_ekleri", "TEXT"),
        ("harcama_bildirimleri", "TEXT"),
        ("gorev_tarih", "TEXT"),
        ("gorev_tarih_iso", "TEXT"),
        ("last_step", "INTEGER DEFAULT 0"),
        ("gorev_il", "TEXT"),
        ("gorev_ilce", "TEXT"),
        ("gorev_firma", "TEXT"),
        ("assigned_to_user_id", "INTEGER"),
        ("assigned_by_user_id", "INTEGER"),
        ("assigned_at", "TEXT"),
    ):
        if column not in existing_columns:
            connection.execute(f"ALTER TABLE forms ADD COLUMN {column} {definition}")

    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_forms_assigned_to ON forms(assigned_to_user_id)"
    )



def determine_form_status(form_data: Dict[str, Any]) -> FormStatus:
    """Formun tamamlanma durumunu belirle."""

    required_fields = [
        "yola_cikis_tarih",
        "yola_cikis_saat",
        "calisma_baslangic_tarih",
        "calisma_baslangic_saat",
        "calisma_bitis_tarih",
        "calisma_bitis_saat",
        "donus_tarih",
        "donus_saat",
    ]

    missing_fields: List[str] = []
    for key in required_fields:
        value = (form_data.get(key) or "").strip()
        if not value:
            missing_fields.append(key)

    status_code = "TAMAMLANDI" if not missing_fields else "YARIM"
    return FormStatus(code=status_code, missing_fields=missing_fields)


def _to_iso_date(value: str | None) -> str | None:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _normalize_for_search(value: str | None) -> str:
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    folded = cleaned.casefold()
    normalized = unicodedata.normalize("NFKD", folded)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def _prepare_payload(form_no: str, form_data: Dict[str, Any], status: FormStatus) -> OrderedDict[str, Any]:
    payload: "OrderedDict[str, Any]" = OrderedDict()
    payload["form_no"] = form_no

    tarih = (form_data.get("tarih") or "").strip()
    gorev_yeri = (form_data.get("gorev_yeri") or "").strip()

    payload["tarih"] = tarih
    payload["tarih_iso"] = _to_iso_date(tarih)
    payload["dok_no"] = (form_data.get("dok_no") or "").strip()
    payload["rev_no"] = (form_data.get("rev_no") or "").strip()
    payload["avans"] = (form_data.get("avans") or "").strip()
    payload["taseron"] = (form_data.get("taseron") or "").strip()
    payload["gorev_tanimi"] = (form_data.get("gorev_tanimi") or "").strip()
    payload["gorev_yeri"] = gorev_yeri
    payload["gorev_yeri_lower"] = _normalize_for_search(gorev_yeri)
    payload["gorev_il"] = (form_data.get("gorev_il") or "").strip()
    payload["gorev_ilce"] = (form_data.get("gorev_ilce") or "").strip()
    payload["gorev_firma"] = (form_data.get("gorev_firma") or "").strip()
    gorev_tarih = (form_data.get("gorev_tarih") or "").strip()
    payload["gorev_tarih"] = gorev_tarih
    payload["gorev_tarih_iso"] = _to_iso_date(gorev_tarih)
    payload["yapilan_isler"] = (form_data.get("yapilan_isler") or "").strip()
    payload["last_step"] = _normalize_last_step(form_data.get("last_step"))

    attachments_raw = form_data.get("gorev_ekleri", [])
    attachments: List[Dict[str, str]] = []
    if isinstance(attachments_raw, str):
        try:
            parsed = json.loads(attachments_raw)
        except json.JSONDecodeError:
            parsed = []
    else:
        parsed = attachments_raw
    if isinstance(parsed, list):
        for item in parsed:
            if isinstance(item, dict) and "filename" in item:
                attachments.append(
                    {
                        "filename": item["filename"],
                        "original_name": item.get("original_name") or item["filename"],
                    }
                )
    payload["gorev_ekleri"] = json.dumps(attachments, ensure_ascii=False)

    expenses_raw = form_data.get("harcama_bildirimleri", [])
    expenses: List[Dict[str, Any]] = []
    if isinstance(expenses_raw, str):
        try:
            parsed_expenses = json.loads(expenses_raw)
        except json.JSONDecodeError:
            parsed_expenses = []
    else:
        parsed_expenses = expenses_raw
    if isinstance(parsed_expenses, list):
        for item in parsed_expenses:
            if not isinstance(item, dict):
                continue
            description = (item.get("description") or "").strip()
            attachments_list: List[Dict[str, str]] = []
            raw_attachments = item.get("attachments", [])
            if isinstance(raw_attachments, list):
                for attachment in raw_attachments:
                    if isinstance(attachment, dict) and attachment.get("filename"):
                        attachments_list.append(
                            {
                                "filename": attachment["filename"],
                                "original_name": attachment.get("original_name")
                                or attachment["filename"],
                            }
                        )
            expenses.append({"description": description, "attachments": attachments_list})
    payload["harcama_bildirimleri"] = json.dumps(expenses, ensure_ascii=False)

    for key in (
        "yola_cikis_tarih",
        "donus_tarih",
        "calisma_baslangic_tarih",
        "calisma_bitis_tarih",
    ):
        value = (form_data.get(key) or "").strip()
        payload[key] = value
        payload[f"{key}_iso"] = _to_iso_date(value)

    payload["yola_cikis_saat"] = (form_data.get("yola_cikis_saat") or "").strip()
    payload["donus_saat"] = (form_data.get("donus_saat") or "").strip()
    payload["calisma_baslangic_saat"] = (form_data.get("calisma_baslangic_saat") or "").strip()
    payload["calisma_bitis_saat"] = (form_data.get("calisma_bitis_saat") or "").strip()
    payload["mola_suresi"] = (form_data.get("mola_suresi") or "").strip()
    payload["arac_plaka"] = (form_data.get("arac_plaka") or "").strip()
    payload["hazirlayan"] = (form_data.get("hazirlayan") or "").strip()
    payload["durum"] = status.code

    payload["assigned_to_user_id"] = _normalize_optional_int(
        form_data.get("assigned_to_user_id")
    )
    payload["assigned_by_user_id"] = _normalize_optional_int(
        form_data.get("assigned_by_user_id")
    )
    payload["assigned_at"] = (form_data.get("assigned_at") or "").strip() or None

    personel_values = []
    for field in PERSONEL_FIELDS:
        value = (form_data.get(field) or "").strip()
        payload[field] = value
        if value:
            personel_values.append(_normalize_for_search(value))

    payload["personel_search"] = ",".join(personel_values)
    return payload


def _normalize_last_step(value: Any) -> int:
    try:
        numeric = int(value)
    except (TypeError, ValueError):
        return 0
    return max(0, numeric)


def _normalize_optional_int(value: Any) -> Optional[int]:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _persist_form(
    form_no: str,
    form_data: Dict[str, Any],
    status: FormStatus,
    base_path: str = ".",
) -> str:
    payload = _prepare_payload(form_no, form_data, status)

    placeholders = ", ".join(["?"] * len(payload))
    columns = ", ".join(payload.keys())
    updates = ", ".join(f"{column}=excluded.{column}" for column in payload.keys() if column != "form_no")

    with _connect(base_path) as connection:
        connection.execute(
            f"""
            INSERT INTO forms ({columns}, created_at, updated_at)
            VALUES ({placeholders}, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT(form_no) DO UPDATE SET
                {updates},
                updated_at=CURRENT_TIMESTAMP
            """,
            tuple(payload.values()),
        )
        try:
            numeric_form_no = int(form_no)
        except ValueError:
            numeric_form_no = None
        if numeric_form_no is not None:
            connection.execute(
                "UPDATE form_sequence SET last_no = MAX(last_no, ?) WHERE id = 1",
                (numeric_form_no,),
            )
        connection.commit()

    return get_db_path(base_path)


def get_next_form_no(base_path: str = ".") -> str:
    """Veritabanındaki en yüksek form numarasını baz alarak bir sonrakini döndür."""

    with _connect(base_path) as connection:
        row = connection.execute(
            "SELECT last_no FROM form_sequence WHERE id = 1"
        ).fetchone()
        last_no = int(row["last_no"] or 0)
        next_no = last_no + 1
        connection.execute(
            "UPDATE form_sequence SET last_no = ? WHERE id = 1",
            (next_no,),
        )
        connection.commit()
    return str(next_no).zfill(5)


def load_form_data(form_no: str, base_path: str = ".") -> Dict[str, Any]:
    """Veritabanından form verisini iç sözlük olarak döndür."""

    with _connect(base_path) as connection:
        row = connection.execute(
            "SELECT * FROM forms WHERE form_no = ?", (form_no,)
        ).fetchone()

    if row is None:
        raise FormServiceError(f"Form {form_no} bulunamadı.")

    row_keys = row.keys()
    last_step_value = row["last_step"] if "last_step" in row_keys else 0

    form_data: Dict[str, Any] = {
        "form_no": form_no,
        "tarih": row["tarih"] or "",
        "dok_no": row["dok_no"] or "",
        "rev_no": row["rev_no"] or "",
        "avans": row["avans"] or "",
        "taseron": row["taseron"] or "",
        "gorev_tanimi": row["gorev_tanimi"] or "",
        "gorev_yeri": row["gorev_yeri"] or "",
        "gorev_il": row["gorev_il"] or "",
        "gorev_ilce": row["gorev_ilce"] or "",
        "gorev_firma": row["gorev_firma"] or "",
        "gorev_tarih": row["gorev_tarih"] or "",
        "yapilan_isler": row["yapilan_isler"] or "",
        "arac_plaka": row["arac_plaka"] or "",
        "hazirlayan": row["hazirlayan"] or "",
        "durum": (row["durum"] or "YARIM").upper(),
        "mola_suresi": row["mola_suresi"] or "",
        "last_step": _normalize_last_step(last_step_value),
        "assigned_to_user_id": row["assigned_to_user_id"] if "assigned_to_user_id" in row_keys else None,
        "assigned_by_user_id": row["assigned_by_user_id"] if "assigned_by_user_id" in row_keys else None,
        "assigned_at": row["assigned_at"] if "assigned_at" in row_keys else None,
    }

    attachments_raw = row["gorev_ekleri"] or "[]"
    try:
        parsed = json.loads(attachments_raw)
    except (TypeError, json.JSONDecodeError):
        parsed = []
    attachments: List[Dict[str, str]] = []
    if isinstance(parsed, list):
        for item in parsed:
            if isinstance(item, dict) and "filename" in item:
                attachments.append(
                    {
                        "filename": item["filename"],
                        "original_name": item.get("original_name") or item["filename"],
                    }
                )
    form_data["gorev_ekleri"] = attachments

    expenses_raw = row["harcama_bildirimleri"] or "[]"
    try:
        parsed_expenses = json.loads(expenses_raw)
    except (TypeError, json.JSONDecodeError):
        parsed_expenses = []
    expenses: List[Dict[str, Any]] = []
    if isinstance(parsed_expenses, list):
        for item in parsed_expenses:
            if not isinstance(item, dict):
                continue
            description = item.get("description") or ""
            attachments_list: List[Dict[str, str]] = []
            raw_attachments = item.get("attachments", [])
            if isinstance(raw_attachments, list):
                for attachment in raw_attachments:
                    if isinstance(attachment, dict) and attachment.get("filename"):
                        attachments_list.append(
                            {
                                "filename": attachment["filename"],
                                "original_name": attachment.get("original_name")
                                or attachment["filename"],
                            }
                        )
            expenses.append({"description": description, "attachments": attachments_list})
    form_data["harcama_bildirimleri"] = expenses

    for field in PERSONEL_FIELDS:
        form_data[field] = row[field] or ""

    for key in (
        "yola_cikis_tarih",
        "yola_cikis_saat",
        "donus_tarih",
        "donus_saat",
        "calisma_baslangic_tarih",
        "calisma_baslangic_saat",
        "calisma_bitis_tarih",
        "calisma_bitis_saat",
    ):
        form_data[key] = row[key] or ""

    return form_data


def save_partial_form(
    form_no: str,
    form_data: Dict[str, Any],
    base_path: str = ".",
) -> Tuple[str, FormStatus]:
    """Formu kısmi olarak kaydet."""

    status = FormStatus(code="YARIM", missing_fields=[])
    db_path = _persist_form(form_no, form_data, status, base_path=base_path)
    return db_path, status


def assign_form(
    form_no: str,
    *,
    assigned_to_user_id: Optional[int],
    assigned_by_user_id: Optional[int],
    base_path: str = ".",
) -> Optional[str]:
    """Bir formu belirli bir kullanıcıya atar veya atamayı kaldırır."""

    assigned_to = _normalize_optional_int(assigned_to_user_id)
    assigned_by = _normalize_optional_int(assigned_by_user_id)
    assigned_at = datetime.utcnow().isoformat(timespec="seconds") if assigned_to else None

    with _connect(base_path) as connection:
        result = connection.execute(
            "UPDATE forms SET assigned_to_user_id = ?, assigned_by_user_id = ?, assigned_at = ?, updated_at = CURRENT_TIMESTAMP WHERE form_no = ?",
            (assigned_to, assigned_by, assigned_at, form_no),
        )
        if result.rowcount == 0:
            raise FormServiceError(f"Form {form_no} bulunamadı.")
        connection.commit()

    return assigned_at


def save_form(
    form_no: str,
    form_data: Dict[str, Any],
    base_path: str = ".",
) -> Tuple[str, FormStatus]:
    """Formu tamamlanmış veya yarım olarak kaydet."""

    status = determine_form_status(form_data)
    db_path = _persist_form(form_no, form_data, status, base_path=base_path)
    return db_path, status


def list_form_numbers(base_path: str = ".") -> List[str]:
    """Veritabanındaki form numaralarını son oluşturulandan başlayarak döndür."""

    with _connect(base_path) as connection:
        rows = connection.execute(
            "SELECT form_no FROM forms ORDER BY CAST(form_no AS INTEGER) DESC"
        ).fetchall()
    return [row["form_no"] for row in rows]


def search_forms(
    *,
    person: str = "",
    location: str = "",
    start_date: str = "",
    end_date: str = "",
    base_path: str = ".",
) -> List[Dict[str, Any]]:
    """Verilen filtrelere göre form kayıtlarını listele."""

    filters: List[str] = []
    params: List[Any] = []

    person = _normalize_for_search(person)
    location = _normalize_for_search(location)
    start_iso = _to_iso_date(start_date)
    end_iso = _to_iso_date(end_date)

    if person:
        filters.append("personel_search LIKE ?")
        params.append(f"%{person}%")

    if location:
        filters.append("gorev_yeri_lower LIKE ?")
        params.append(f"%{location}%")

    if start_iso:
        filters.append("yola_cikis_tarih_iso IS NOT NULL AND yola_cikis_tarih_iso >= ?")
        params.append(start_iso)

    if end_iso:
        filters.append("yola_cikis_tarih_iso IS NOT NULL AND yola_cikis_tarih_iso <= ?")
        params.append(end_iso)

    where_clause = ""
    if filters:
        where_clause = " WHERE " + " AND ".join(filters)

    query = (
        "SELECT form_no, tarih, gorev_yeri, hazirlayan, durum, "
        "yola_cikis_tarih, yola_cikis_tarih_iso, gorev_tanimi, avans, taseron,"
        + ", ".join(PERSONEL_FIELDS)
        + " FROM forms"
        + where_clause
        + " ORDER BY COALESCE(yola_cikis_tarih_iso, '') DESC, CAST(form_no AS INTEGER) DESC"
    )

    with _connect(base_path) as connection:
        rows = connection.execute(query, tuple(params)).fetchall()

    results: List[Dict[str, Any]] = []
    for row in rows:
        personel = [row[field] for field in PERSONEL_FIELDS if row[field]]
        results.append(
            {
                "form_no": row["form_no"],
                "tarih": row["tarih"] or "",
                "gorev_yeri": row["gorev_yeri"] or "",
                "hazirlayan": row["hazirlayan"] or "",
                "durum": row["durum"] or "",
                "yola_cikis_tarih": row["yola_cikis_tarih"] or "",
                "yola_cikis_tarih_iso": row["yola_cikis_tarih_iso"],
                "personel": personel,
                "gorev_tanimi": row["gorev_tanimi"] or "",
                "avans": row["avans"] or "",
                "taseron": row["taseron"] or "",
            }
        )

    return results


def list_forms_for_assignee(
    assigned_user_id: int,
    *,
    base_path: str = ".",
) -> List[Dict[str, Any]]:
    """Belirli bir çalışana atanan formları döndür."""

    query = (
        "SELECT f.form_no, f.gorev_yeri, f.gorev_tanimi, f.durum, f.gorev_tarih, "
        "f.yola_cikis_tarih, f.assigned_at, f.assigned_by_user_id, u.full_name AS assigned_by_name, "
        "f.updated_at "
        "FROM forms f "
        "LEFT JOIN users u ON u.id = f.assigned_by_user_id "
        "WHERE f.assigned_to_user_id = ? "
        "ORDER BY f.updated_at DESC"
    )

    with _connect(base_path) as connection:
        rows = connection.execute(query, (assigned_user_id,)).fetchall()

    assignments: List[Dict[str, Any]] = []
    for row in rows:
        assignments.append(
            {
                "form_no": row["form_no"],
                "gorev_yeri": row["gorev_yeri"] or "",
                "gorev_tanimi": row["gorev_tanimi"] or "",
                "durum": (row["durum"] or "").upper(),
                "gorev_tarih": row["gorev_tarih"] or row["yola_cikis_tarih"] or "",
                "assigned_at": row["assigned_at"],
                "assigned_by_user_id": row["assigned_by_user_id"],
                "assigned_by_name": row["assigned_by_name"] or "",
                "updated_at": row["updated_at"],
            }
        )

    return assignments


def get_reporting_summary(
    *,
    start_date: str = "",
    end_date: str = "",
    base_path: str = ".",
) -> Dict[str, Any]:
    """Derlenmiş raporlama metriklerini döndür."""

    start_iso = _to_iso_date(start_date)
    end_iso = _to_iso_date(end_date)

    filters: List[str] = []
    params: List[Any] = []

    if start_iso:
        filters.append(
            "COALESCE(yola_cikis_tarih_iso, gorev_tarih_iso) IS NOT NULL "
            "AND COALESCE(yola_cikis_tarih_iso, gorev_tarih_iso) >= ?"
        )
        params.append(start_iso)

    if end_iso:
        filters.append(
            "COALESCE(yola_cikis_tarih_iso, gorev_tarih_iso) IS NOT NULL "
            "AND COALESCE(yola_cikis_tarih_iso, gorev_tarih_iso) <= ?"
        )
        params.append(end_iso)

    where_clause = ""
    if filters:
        where_clause = " WHERE " + " AND ".join(filters)

    columns = [
        "form_no",
        "gorev_tanimi",
        "avans",
        "harcama_bildirimleri",
        "yola_cikis_tarih",
        "yola_cikis_tarih_iso",
        "yola_cikis_saat",
        "donus_tarih",
        "donus_tarih_iso",
        "donus_saat",
        "calisma_baslangic_tarih",
        "calisma_baslangic_tarih_iso",
        "calisma_baslangic_saat",
        "calisma_bitis_tarih",
        "calisma_bitis_tarih_iso",
        "calisma_bitis_saat",
        "gorev_yeri",
        "gorev_il",
        "gorev_ilce",
        "gorev_firma",
        "gorev_tarih",
        "gorev_tarih_iso",
    ]
    columns.extend(PERSONEL_FIELDS)

    query = (
        "SELECT "
        + ", ".join(columns)
        + " FROM forms"
        + where_clause
        + " ORDER BY COALESCE(yola_cikis_tarih_iso, gorev_tarih_iso, '') DESC, CAST(form_no AS INTEGER) DESC"
    )

    with _connect(base_path) as connection:
        rows = connection.execute(query, tuple(params)).fetchall()

    def _combine_datetime(row: sqlite3.Row, date_key: str, iso_key: str, time_key: str) -> datetime | None:
        date_iso = row[iso_key] or None
        if not date_iso:
            date_iso = _to_iso_date(row[date_key])
        if not date_iso:
            return None
        time_value = (row[time_key] or "").strip()
        if not time_value:
            time_value = "00:00"
        if len(time_value.split(":")) == 2:
            time_value = f"{time_value}:00"
        try:
            return datetime.fromisoformat(f"{date_iso}T{time_value}")
        except ValueError:
            return None

    person_counts: Dict[str, int] = {}
    unique_people: set[str] = set()
    total_travel_hours = 0.0
    total_work_hours = 0.0
    travel_samples = 0
    work_samples = 0
    expense_labels: List[str] = []
    expense_values: List[float] = []
    location_counter: Dict[Tuple[str, str, str], int] = {}
    forms_summary: List[Dict[str, Any]] = []

    for row in rows:
        personel = [row[field] for field in PERSONEL_FIELDS if row[field]]
        for person in personel:
            normalized = person.strip()
            if not normalized:
                continue
            person_counts[normalized] = person_counts.get(normalized, 0) + 1
            unique_people.add(normalized)

        start_dt = _combine_datetime(row, "yola_cikis_tarih", "yola_cikis_tarih_iso", "yola_cikis_saat")
        end_dt = _combine_datetime(row, "donus_tarih", "donus_tarih_iso", "donus_saat")
        travel_hours = None
        if start_dt and end_dt and end_dt >= start_dt:
            delta = end_dt - start_dt
            travel_hours = round(delta.total_seconds() / 3600, 2)
            total_travel_hours += travel_hours
            travel_samples += 1

        work_start = _combine_datetime(
            row,
            "calisma_baslangic_tarih",
            "calisma_baslangic_tarih_iso",
            "calisma_baslangic_saat",
        )
        work_end = _combine_datetime(
            row,
            "calisma_bitis_tarih",
            "calisma_bitis_tarih_iso",
            "calisma_bitis_saat",
        )
        work_hours = None
        if work_start and work_end and work_end >= work_start:
            delta = work_end - work_start
            work_hours = round(delta.total_seconds() / 3600, 2)
            total_work_hours += work_hours
            work_samples += 1

        expenses_raw = row["harcama_bildirimleri"] or "[]"
        try:
            expenses = json.loads(expenses_raw)
        except (TypeError, json.JSONDecodeError):
            expenses = []
        if not isinstance(expenses, list):
            expenses = []
        expense_count = len(expenses)
        expense_labels.append(row["form_no"])
        expense_values.append(float(expense_count))

        location_key = (
            (row["gorev_il"] or "").strip(),
            (row["gorev_ilce"] or "").strip(),
            (row["gorev_firma"] or "").strip(),
        )
        if not any(location_key):
            location_key = ((row["gorev_yeri"] or "Belirtilmedi").strip(), "", "")
        location_counter[location_key] = location_counter.get(location_key, 0) + 1

        forms_summary.append(
            {
                "form_no": row["form_no"],
                "gorev_tanimi": row["gorev_tanimi"] or "",
                "personel": personel,
                "travel_hours": travel_hours,
                "work_hours": work_hours,
                "expense_count": expense_count,
                "gorev_il": row["gorev_il"] or "",
                "gorev_ilce": row["gorev_ilce"] or "",
                "gorev_firma": row["gorev_firma"] or "",
                "gorev_yeri": row["gorev_yeri"] or "",
                "gorev_tarih": row["gorev_tarih"] or "",
            }
        )

    sorted_persons = sorted(person_counts.items(), key=lambda item: (-item[1], item[0]))
    person_breakdown = [
        {"person": name, "count": count}
        for name, count in sorted_persons
    ]

    sorted_locations = sorted(
        location_counter.items(),
        key=lambda item: (-item[1], item[0]),
    )
    location_breakdown = [
        {
            "label": ", ".join(filter(None, key)).strip() or "Belirtilmedi",
            "count": count,
        }
        for key, count in sorted_locations
    ]

    return {
        "total_forms": len(forms_summary),
        "unique_person_count": len(unique_people),
        "person_breakdown": person_breakdown,
        "forms": forms_summary,
        "travel_hours": {
            "total": round(total_travel_hours, 2),
            "average": round(total_travel_hours / travel_samples, 2) if travel_samples else 0.0,
            "samples": travel_samples,
        },
        "work_hours": {
            "total": round(total_work_hours, 2),
            "average": round(total_work_hours / work_samples, 2) if work_samples else 0.0,
            "samples": work_samples,
        },
        "expense_chart": {
            "labels": expense_labels,
            "values": expense_values,
        },
        "locations": location_breakdown,
        "filters": {
            "start_date": start_iso or "",
            "end_date": end_iso or "",
        },
    }


def _build_excel_workbook(form_no: str, form_data: Dict[str, Any], status: FormStatus) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Görev Formu"

    header_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    status_fill = (
        PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        if status.is_complete
        else PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    )
    status_value_fill = (
        PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")
        if status.is_complete
        else PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row_index = 1
    worksheet[f"A{row_index}"] = "DELTA PROJE - GÖREV FORMU"
    worksheet[f"A{row_index}"].font = Font(size=16, bold=True, color="D32F2F")
    worksheet.merge_cells(f"A{row_index}:B{row_index}")
    row_index += 1

    data_map: Sequence[Tuple[str, str]] = [
        ("Form No", form_no),
        ("Tarih", form_data.get("tarih", "")),
        ("Görev Tarihi", form_data.get("gorev_tarih", "")),
        ("DOK.NO", form_data.get("dok_no", "")),
        ("REV.NO/TRH", form_data.get("rev_no", "")),
        ("", ""),
        ("Görevli Personel", ""),
    ]

    for label, value in data_map:
        if label:
            worksheet[f"A{row_index}"] = label
            worksheet[f"A{row_index}"].font = Font(bold=True)
            worksheet[f"A{row_index}"].fill = header_fill
            worksheet[f"A{row_index}"].border = border
            worksheet[f"B{row_index}"] = value
            worksheet[f"B{row_index}"].border = border
        row_index += 1

    for field in PERSONEL_FIELDS:
        worksheet[f"A{row_index}"] = field.replace("_", " ").title()
        worksheet[f"A{row_index}"].border = border
        worksheet[f"B{row_index}"] = form_data.get(field, "")
        worksheet[f"B{row_index}"].border = border
        row_index += 1

    row_index += 1

    def format_datetime(date_key: str, time_key: str) -> str:
        tarih = (form_data.get(date_key) or "").strip()
        saat = (form_data.get(time_key) or "").strip()
        if tarih and saat:
            return f"{tarih} {saat}"
        return tarih or saat

    mola = (form_data.get("mola_suresi") or "").strip()
    mola_text = f"{mola} dakika" if mola else ""
    yapilan_isler = (form_data.get("yapilan_isler") or "").strip()
    attachment_names = []
    for item in form_data.get("gorev_ekleri", []):
        if isinstance(item, dict):
            original = item.get("original_name")
            if original:
                attachment_names.append(original)
    attachments_text = "\n".join(attachment_names)

    expense_lines: List[str] = []
    for index, expense in enumerate(form_data.get("harcama_bildirimleri", []), 1):
        if not isinstance(expense, dict):
            continue
        description = (expense.get("description") or "Açıklama belirtilmedi").strip() or "Açıklama belirtilmedi"
        receipt_names: List[str] = []
        for receipt in expense.get("attachments", []) or []:
            if isinstance(receipt, dict):
                name = receipt.get("original_name") or receipt.get("filename")
                if name:
                    receipt_names.append(name)
        if receipt_names:
            expense_lines.append(f"{index}. {description} (Ekler: {', '.join(receipt_names)})")
        else:
            expense_lines.append(f"{index}. {description}")
    expenses_text = "\n".join(expense_lines)

    detail_map: Sequence[Tuple[str, str]] = [
        ("Avans Tutarı", form_data.get("avans", "")),
        ("Taşeron Şirket", form_data.get("taseron", "")),
        ("Görevin Tanımı", form_data.get("gorev_tanimi", "")),
        ("Görev Yeri", form_data.get("gorev_yeri", "")),
        ("Görev İli", form_data.get("gorev_il", "")),
        ("Görev İlçesi", form_data.get("gorev_ilce", "")),
        ("Firma/Lokasyon", form_data.get("gorev_firma", "")),
        ("Yapılan İşler", yapilan_isler),
        ("Harcama Bildirimleri", expenses_text),
        ("Ekler", attachments_text),
        ("", ""),
        ("Yola Çıkış", format_datetime("yola_cikis_tarih", "yola_cikis_saat")),
        ("Dönüş", format_datetime("donus_tarih", "donus_saat")),
        (
            "Çalışma Başlangıç",
            format_datetime("calisma_baslangic_tarih", "calisma_baslangic_saat"),
        ),
        (
            "Çalışma Bitiş",
            format_datetime("calisma_bitis_tarih", "calisma_bitis_saat"),
        ),
        ("Toplam Mola", mola_text),
        ("", ""),
        ("Araç Plaka No", form_data.get("arac_plaka", "")),
        ("Hazırlayan", form_data.get("hazirlayan", "")),
        ("", ""),
        ("DURUM", status.code),
    ]

    for label, value in detail_map:
        if label:
            worksheet[f"A{row_index}"] = label
            worksheet[f"A{row_index}"].font = Font(bold=True)
            if label == "DURUM":
                worksheet[f"A{row_index}"].fill = status_fill
                worksheet[f"B{row_index}"].fill = status_value_fill
            else:
                worksheet[f"A{row_index}"].fill = header_fill
            worksheet[f"A{row_index}"].border = border
            worksheet[f"B{row_index}"] = value
            worksheet[f"B{row_index}"].border = border
        row_index += 1

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 60
    return workbook


def export_form_to_excel(
    form_no: str,
    form_data: Dict[str, Any],
) -> io.BytesIO:
    """Formu Excel dosyası olarak dışa aktar."""

    status = determine_form_status(form_data)
    workbook = _build_excel_workbook(form_no, form_data, status)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_form_to_pdf(
    form_no: str,
    form_data: Dict[str, Any],
) -> io.BytesIO:
    """Formu PDF dosyası olarak dışa aktar."""

    status = determine_form_status(form_data)
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    margin = 2 * cm
    y_position = height - margin

    pdf.setFillColor(colors.HexColor("#D32F2F"))
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(margin, y_position, "DELTA PROJE - GÖREV FORMU")
    y_position -= 1.2 * cm

    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica-Bold", 10)
    metadata = [
        ("Form No", form_no),
        ("Tarih", form_data.get("tarih", "")),
        ("Görev Tarihi", form_data.get("gorev_tarih", "")),
        ("DOK.NO", form_data.get("dok_no", "")),
        ("REV.NO/TRH", form_data.get("rev_no", "")),
    ]
    for label, value in metadata:
        pdf.drawString(margin, y_position, f"{label}: ")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(margin + 4.5 * cm, y_position, value or "-")
        pdf.setFont("Helvetica-Bold", 10)
        y_position -= 0.7 * cm

    y_position -= 0.3 * cm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColor(colors.HexColor("#0D47A1"))
    pdf.drawString(margin, y_position, "Görevli Personel")
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica", 10)
    y_position -= 0.6 * cm
    for field in PERSONEL_FIELDS:
        value = form_data.get(field, "") or "-"
        pdf.drawString(margin + 0.5 * cm, y_position, f"{field.replace('_', ' ').title()}: {value}")
        y_position -= 0.5 * cm

    y_position -= 0.2 * cm
    yapilan_isler = (form_data.get("yapilan_isler") or "").strip()
    attachment_names = []
    for item in form_data.get("gorev_ekleri", []):
        if isinstance(item, dict):
            original = item.get("original_name")
            if original:
                attachment_names.append(original)
    attachments_text = ", ".join(attachment_names)

    expense_lines: List[str] = []
    for index, expense in enumerate(form_data.get("harcama_bildirimleri", []), 1):
        if not isinstance(expense, dict):
            continue
        description = (expense.get("description") or "Açıklama belirtilmedi").strip() or "Açıklama belirtilmedi"
        receipt_names: List[str] = []
        for receipt in expense.get("attachments", []) or []:
            if isinstance(receipt, dict):
                name = receipt.get("original_name") or receipt.get("filename")
                if name:
                    receipt_names.append(name)
        if receipt_names:
            expense_lines.append(f"{index}. {description} (Ekler: {', '.join(receipt_names)})")
        else:
            expense_lines.append(f"{index}. {description}")
    expenses_text = "; ".join(expense_lines)

    sections: Iterable[Tuple[str, str]] = (
        ("Avans Tutarı", form_data.get("avans", "")),
        ("Taşeron Şirket", form_data.get("taseron", "")),
        ("Görevin Tanımı", form_data.get("gorev_tanimi", "")),
        ("Görev Yeri", form_data.get("gorev_yeri", "")),
        ("Görev İli", form_data.get("gorev_il", "")),
        ("Görev İlçesi", form_data.get("gorev_ilce", "")),
        ("Firma/Lokasyon", form_data.get("gorev_firma", "")),
        ("Yapılan İşler", yapilan_isler),
        ("Harcama Bildirimleri", expenses_text),
        ("Ekler", attachments_text),
        ("Yola Çıkış", f"{form_data.get('yola_cikis_tarih', '')} {form_data.get('yola_cikis_saat', '')}".strip()),
        ("Dönüş", f"{form_data.get('donus_tarih', '')} {form_data.get('donus_saat', '')}".strip()),
        (
            "Çalışma Başlangıç",
            f"{form_data.get('calisma_baslangic_tarih', '')} {form_data.get('calisma_baslangic_saat', '')}".strip(),
        ),
        (
            "Çalışma Bitiş",
            f"{form_data.get('calisma_bitis_tarih', '')} {form_data.get('calisma_bitis_saat', '')}".strip(),
        ),
        ("Toplam Mola", (form_data.get("mola_suresi") or "") + (" dakika" if form_data.get("mola_suresi") else "")),
        ("Araç Plaka No", form_data.get("arac_plaka", "")),
        ("Hazırlayan", form_data.get("hazirlayan", "")),
    )

    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColor(colors.HexColor("#0D47A1"))
    pdf.drawString(margin, y_position, "Görev Bilgileri")
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica", 10)
    y_position -= 0.6 * cm

    for label, value in sections:
        pdf.drawString(margin + 0.5 * cm, y_position, f"{label}: {value or '-'}")
        y_position -= 0.5 * cm

    if y_position < margin + 4 * cm:
        pdf.showPage()
        y_position = height - margin

    pdf.setFont("Helvetica-Bold", 12)
    pdf.setFillColor(colors.HexColor("#4CAF50" if status.is_complete else "#FF9800"))
    pdf.drawString(margin, y_position, f"DURUM: {status.code}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer


__all__ = [
    "DB_FILENAME",
    "FormServiceError",
    "FormStatus",
    "determine_form_status",
    "export_form_to_excel",
    "export_form_to_pdf",
    "get_db_path",
    "get_next_form_no",
    "get_reporting_summary",
    "list_form_numbers",
    "load_form_data",
    "save_form",
    "save_partial_form",
    "search_forms",
]
